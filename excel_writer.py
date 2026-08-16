import os
import glob
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADERS = [
    "№",
    "Цех",
    "Организация",
    "Дата",
    "НАЛ или БЕЗНАЛ",
    "Карта",
    "Цена",
    "Количество",
    "Сумма",
    "Наименование ТМЦ",
    "Наименование товара",
    "Поставщик",
    "Исполнитель",
    "Примечание",
    "Фото чека"
]

COLUMN_WIDTHS = {
    "A": 5,   
    "B": 15,  
    "C": 20,  
    "D": 22,  
    "E": 20,  
    "F": 22,  
    "G": 15,  
    "H": 15,  
    "I": 15,  
    "J": 25,  
    "K": 25,  
    "L": 20,  
    "M": 20,  
    "N": 25,  
    "O": 30,  
}

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
os.makedirs(DATA_DIR, exist_ok=True)

def get_filename(date_obj=None):
    if not date_obj:
        date_obj = datetime.now()
    return os.path.join(DATA_DIR, f"data_{date_obj.strftime('%m_%Y')}.xlsx")

def get_all_files():
    return sorted(glob.glob(os.path.join(DATA_DIR, "data_*.xlsx")))

def apply_header_styles(ws):
    
    header_font = Font(name="Calibri", size=12, bold=True, color="000000")
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") 
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    ws.append(HEADERS)
    for col in range(1, len(HEADERS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
    for col_letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
        
    ws.freeze_panes = "A2"
    
    ws.auto_filter.ref = f"A1:O1"

def ensure_sheet_exists(wb, shop_name):
    
    safe_shop_name = shop_name[:31]
    
    if 'Sheet' in wb.sheetnames and len(wb.sheetnames) == 1 and wb['Sheet'].max_row == 1:
        wb['Sheet'].title = safe_shop_name
        ws = wb[safe_shop_name]
        apply_header_styles(ws)
        return ws
    elif safe_shop_name not in wb.sheetnames:
        ws = wb.create_sheet(title=safe_shop_name)
        apply_header_styles(ws)
        return ws
    else:
        return wb[safe_shop_name]

def get_last_date(ws):
    for row in range(ws.max_row, 1, -1):
        val = ws.cell(row=row, column=4).value
        if val:
            if isinstance(val, str) and len(val) >= 10:
                return val[:10]
    return None

def get_next_number(ws):
    for row in range(ws.max_row, 1, -1):
        val = ws.cell(row=row, column=1).value
        if isinstance(val, (int, float)):
            return int(val) + 1
        elif isinstance(val, str) and val.isdigit():
            return int(val) + 1
    return 1

def clean_bottom_totals(ws):
    
    while ws.max_row > 1:
        val_b = ws.cell(row=ws.max_row, column=2).value 
        val_d = ws.cell(row=ws.max_row, column=4).value 
        
        is_total = isinstance(val_b, str) and str(val_b).startswith("Итого")
        is_empty = (val_b is None and val_d is None)
        
        if is_total or is_empty:
            ws.delete_rows(ws.max_row)
        else:
            break

def find_day_start_row(ws, end_row, date_str):
    start_row = end_row
    for r in range(end_row, 1, -1):
        val_date = ws.cell(row=r, column=4).value
        if val_date and isinstance(val_date, str) and len(val_date) >= 10:
            if val_date.startswith(date_str):
                start_row = r
            else:
                break 
    return start_row

def add_daily_total(ws, date_str):
    start_row = find_day_start_row(ws, ws.max_row, date_str)
    if start_row > 1:
        end_row = ws.max_row
        row = [""] * 15
        row[1] = f"Итого за {date_str}:"
        
        row[8] = f"=SUBTOTAL(9, I{start_row}:I{end_row})"
        ws.append(row)
        
        curr_row = ws.max_row
        ws.cell(row=curr_row, column=2).font = Font(bold=True)
        sum_cell = ws.cell(row=curr_row, column=9)
        sum_cell.number_format = '#,##0'
        sum_cell.font = Font(bold=True)
        
        fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for c in range(1, 16):
            ws.cell(row=curr_row, column=c).fill = fill

def add_monthly_total(ws):
    end_row = ws.max_row
    row = [""] * 15
    row[1] = "Итого за месяц:"
    row[8] = f"=SUBTOTAL(9, I2:I{end_row})"
    ws.append(row)
    
    curr_row = ws.max_row
    ws.cell(row=curr_row, column=2).font = Font(bold=True, color="FFFFFF")
    sum_cell = ws.cell(row=curr_row, column=9)
    sum_cell.number_format = '#,##0'
    sum_cell.font = Font(bold=True, color="FFFFFF")
    
    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for c in range(1, 16):
        ws.cell(row=curr_row, column=c).fill = fill

def add_record(data: dict):
    now = datetime.now()
    filename = get_filename(now)
    
    if not os.path.exists(filename):
        wb = Workbook()
    else:
        wb = load_workbook(filename)
        
    shop = data.get("shop", "Неизвестно")
    ws = ensure_sheet_exists(wb, shop)
    
    clean_bottom_totals(ws)
    
    receipt_date_str = data.get("receipt_date", "")
    last_data_date = get_last_date(ws)
    
    current_time = now.strftime("%H:%M:%S")
    if receipt_date_str:
        current_datetime_str = f"{receipt_date_str} {current_time}"
        current_date_str = receipt_date_str
    else:
        current_datetime_str = now.strftime("%d.%m.%Y %H:%M:%S")
        current_date_str = now.strftime("%d.%m.%Y")
    
    if last_data_date and last_data_date != current_date_str:
        add_daily_total(ws, last_data_date)
        ws.append([])
        
    payment_type = data.get("payment_type", "")
    card_number = data.get("card_number", "")
    supplier = data.get("supplier", "")
    performer = data.get("performer", "")
    organization = data.get("organization", "")
    photo_path = data.get("photo_path", "")
    
    items = data.get("items", [])
    if not items:
        return
        
    for item in items:
        try:
            p_str = str(item.get("price", "0")).replace(",", ".").replace(" ", "")
            q_str = str(item.get("quantity", "0")).replace(",", ".").replace(" ", "")
            price = float(p_str) if p_str else 0
            qty = float(q_str) if q_str else 0
        except ValueError:
            price = 0
            qty = 0
            
        if photo_path and photo_path.startswith("http"):
            photo_links = [u for u in photo_path.split(",") if u]
            photo_link = f'=HYPERLINK("{photo_links[0]}", "Смотреть")'
            extra_photo_links = [
                f'=HYPERLINK("{u}", "Смотреть")'
                for u in photo_links[1:]
            ]
        elif photo_path:
            photo_links = [photo_path]
            base_dir = os.path.dirname(os.path.abspath(__file__))
            abs_photo = os.path.join(base_dir, "photos", photo_path)
            photo_link = f'=HYPERLINK("file://{abs_photo}", "Смотреть")'
            extra_photo_links = []
        else:
            photo_link = ""
            extra_photo_links = []

        next_num = get_next_number(ws)

        row = [
            next_num,
            shop,
            organization,
            current_datetime_str,
            payment_type,
            card_number,
            price,
            qty,
            None,   
            data.get("tmc_group", ""),
            item.get("nomenclature", ""),
            supplier,
            performer,
            data.get("note", ""),
            photo_link
        ]
        
        ws.append(row)

        current_row = ws.max_row

        for offset, link in enumerate(extra_photo_links, start=1):
            ws.cell(row=current_row, column=15 + offset).value = link

        ws.cell(row=current_row, column=9).value = f"=G{current_row}*H{current_row}"
        ws.cell(row=current_row, column=7).number_format = '#,##0'
        ws.cell(row=current_row, column=8).number_format = '#,##0.##'
        ws.cell(row=current_row, column=9).number_format = '#,##0'
        ws.cell(row=current_row, column=9).font = Font(bold=True)

    add_daily_total(ws, current_date_str)
    
    ws.append([])
    
    add_monthly_total(ws)
    
    wb.save(filename)
